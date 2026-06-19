from __future__ import annotations

import json
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.invoice.service import export_all, list_games, list_invoices, list_payments, list_schools


def _autosize_columns(ws) -> None:  # noqa: ANN001
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)


def _write_sheet(ws, headers: list[str], rows: list[list[Any]]) -> None:  # noqa: ANN001
    bold = Font(bold=True)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = bold
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    _autosize_columns(ws)


def export_json_backup() -> dict[str, Any]:
    payload = export_all()
    payload["exported_at"] = datetime.now(UTC).isoformat()
    return payload


def export_excel_backup() -> bytes:
    wb = Workbook()
    readme = wb.active
    readme.title = "_README"
    readme["A1"] = "Invoice Database Backup"
    readme["A2"] = f"Exported: {datetime.now(UTC).isoformat()}"
    readme["A3"] = "Sheets mirror the normalized schema from Invoice_Database.accdb rebuild."

    schools_ws = wb.create_sheet("schools")
    school_rows = []
    for school in list_schools():
        school_rows.append(
            [
                school.id,
                school.school_name,
                school.address,
                school.city,
                school.state,
                school.zip,
                school.field_location,
                school.subsite,
                school.varsity_game_time,
                school.jv_game_time,
                school.rank,
                "Y" if school.parkway_district else "N",
                ", ".join(school.sports),
            ]
        )
    _write_sheet(
        schools_ws,
        [
            "id",
            "school_name",
            "address",
            "city",
            "state",
            "zip",
            "field_location",
            "subsite",
            "varsity_game_time",
            "jv_game_time",
            "rank",
            "parkway_district",
            "sports",
        ],
        school_rows,
    )

    invoices_ws = wb.create_sheet("invoices")
    invoice_rows = []
    for invoice in list_invoices():
        invoice_rows.append(
            [
                invoice.id,
                invoice.school_name,
                invoice.season_year,
                invoice.sport,
                invoice.base_amount,
                invoice.revision_amount,
                invoice.dual_sport_fee,
                invoice.ranking_services,
                invoice.c_team_scheduling,
                invoice.fh_ranking_services,
                invoice.total_amount,
                invoice.amount_paid,
                invoice.balance_due,
                invoice.address_note,
                invoice.collection_status,
                invoice.notes,
            ]
        )
    _write_sheet(
        invoices_ws,
        [
            "id",
            "school_name",
            "season_year",
            "sport",
            "base_amount",
            "revision_amount",
            "dual_sport_fee",
            "ranking_services",
            "c_team_scheduling",
            "fh_ranking_services",
            "total_amount",
            "amount_paid",
            "balance_due",
            "address_note",
            "collection_status",
            "notes",
        ],
        invoice_rows,
    )

    payments_ws = wb.create_sheet("payments")
    payment_rows = []
    for payment in list_payments():
        payment_rows.append(
            [
                payment.id,
                payment.school_name,
                payment.season_year,
                payment.sport,
                payment.invoice_id,
                payment.amount_paid,
                payment.date_paid,
                payment.payment_method,
                payment.notes,
            ]
        )
    _write_sheet(
        payments_ws,
        [
            "id",
            "school_name",
            "season_year",
            "sport",
            "invoice_id",
            "amount_paid",
            "date_paid",
            "payment_method",
            "notes",
        ],
        payment_rows,
    )

    games_ws = wb.create_sheet("games")
    game_rows = []
    for game in list_games():
        game_rows.append(
            [
                game.id,
                game.game_date,
                game.game_time,
                game.level,
                game.home_school_name,
                game.away_school_name,
                game.field_location,
                game.season_year,
                game.sport,
            ]
        )
    _write_sheet(
        games_ws,
        [
            "id",
            "game_date",
            "game_time",
            "level",
            "home_team",
            "away_team",
            "field_location",
            "season_year",
            "sport",
        ],
        game_rows,
    )

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def write_json_backup(path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(export_json_backup(), indent=2), encoding="utf-8")
    return path


def write_excel_backup(path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(export_excel_backup())
    return path
