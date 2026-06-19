from __future__ import annotations

import re
from collections.abc import Callable, Iterator
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any


def serialize_cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        return ILLEGAL_EXCEL_CHARS.sub("", value)
    if isinstance(value, memoryview):
        value = value.tobytes()
    if isinstance(value, (bytes, bytearray)):
        return ILLEGAL_EXCEL_CHARS.sub("", value.decode("utf-8", errors="replace"))
    if isinstance(value, datetime):
        return value.isoformat()
    return ILLEGAL_EXCEL_CHARS.sub("", str(value))

import psycopg2
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
from openpyxl.workbook.child import INVALID_TITLE_REGEX
from openpyxl.worksheet.worksheet import Worksheet
from psycopg2 import sql
from psycopg2.extensions import connection as PgConnection

from app.connection import connection_host
from app.export.logical_keys import (
    build_mermaid_er,
    build_mermaid_er_sections,
    explicit_logical_edges,
    load_logical_keys_config,
    logical_key_rows,
)
from app.export.schema_diagram import write_schema_diagram_html

DEFAULT_SCHEMAS = ("public", "archive")
BATCH_SIZE = 1000
ILLEGAL_EXCEL_CHARS = re.compile(r"[\000-\010\013\014\016-\037]")
DOC_SHEETS = ("_README", "_Tables", "_Columns", "_ForeignKeys", "_LogicalKeys", "_ER_Diagram")
ProgressCallback = Callable[[str, int, int], None]


@dataclass(frozen=True)
class TableRef:
    schema_name: str
    table_name: str
    table_type: str


@dataclass(frozen=True)
class ExportSummary:
    output_path: str
    diagram_path: str
    table_count: int
    sheet_count: int
    exported_at: datetime


def discover_tables(conn: PgConnection, schemas: tuple[str, ...]) -> list[TableRef]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT table_schema, table_name, table_type
            FROM information_schema.tables
            WHERE table_schema = ANY(%s)
              AND table_type IN ('BASE TABLE', 'VIEW')
            ORDER BY table_schema, table_name
            """,
            (list(schemas),),
        )
        return [
            TableRef(schema_name=row[0], table_name=row[1], table_type=row[2])
            for row in cur.fetchall()
        ]


def fetch_table_inventory(conn: PgConnection, schemas: tuple[str, ...]) -> list[list[Any]]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT
                t.table_schema,
                t.table_name,
                t.table_type,
                COALESCE(c.reltuples::bigint, 0) AS approx_row_count,
                COUNT(col.column_name) AS column_count
            FROM information_schema.tables t
            LEFT JOIN pg_class c
              ON c.relname = t.table_name
             AND c.relnamespace = (
                SELECT oid FROM pg_namespace WHERE nspname = t.table_schema
             )
            LEFT JOIN information_schema.columns col
              ON col.table_schema = t.table_schema
             AND col.table_name = t.table_name
            WHERE t.table_schema = ANY(%s)
              AND t.table_type IN ('BASE TABLE', 'VIEW')
            GROUP BY t.table_schema, t.table_name, t.table_type, c.reltuples
            ORDER BY t.table_schema, t.table_name
            """,
            (list(schemas),),
        )
        rows = cur.fetchall()
    return [
        [schema, name, table_type, int(approx or 0), int(cols or 0)]
        for schema, name, table_type, approx, cols in rows
    ]


def fetch_columns(conn: PgConnection, schemas: tuple[str, ...]) -> list[list[Any]]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT table_schema, table_name, column_name, data_type,
                   is_nullable, column_default
            FROM information_schema.columns
            WHERE table_schema = ANY(%s)
            ORDER BY table_schema, table_name, ordinal_position
            """,
            (list(schemas),),
        )
        return [list(row) for row in cur.fetchall()]


def fetch_foreign_keys(conn: PgConnection, schemas: tuple[str, ...]) -> list[list[Any]]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT
                ns_src.nspname AS table_schema,
                src.relname AS table_name,
                att_src.attname AS column_name,
                ns_tgt.nspname AS foreign_table_schema,
                tgt.relname AS foreign_table_name,
                att_tgt.attname AS foreign_column_name,
                con.conname AS constraint_name
            FROM pg_constraint con
            JOIN pg_class src ON src.oid = con.conrelid
            JOIN pg_namespace ns_src ON ns_src.oid = src.relnamespace
            JOIN pg_class tgt ON tgt.oid = con.confrelid
            JOIN pg_namespace ns_tgt ON ns_tgt.oid = tgt.relnamespace
            JOIN LATERAL unnest(con.conkey) WITH ORDINALITY AS src_cols(attnum, ord) ON TRUE
            JOIN LATERAL unnest(con.confkey) WITH ORDINALITY AS tgt_cols(attnum, ord)
              ON tgt_cols.ord = src_cols.ord
            JOIN pg_attribute att_src
              ON att_src.attrelid = src.oid AND att_src.attnum = src_cols.attnum
            JOIN pg_attribute att_tgt
              ON att_tgt.attrelid = tgt.oid AND att_tgt.attnum = tgt_cols.attnum
            WHERE con.contype = 'f'
              AND ns_src.nspname = ANY(%s)
            ORDER BY ns_src.nspname, src.relname, src_cols.ord
            """,
            (list(schemas),),
        )
        return [list(row) for row in cur.fetchall()]




def sanitize_sheet_name(schema_name: str, table_name: str) -> str:
    raw = f"{schema_name}.{table_name}" if schema_name != "public" else table_name
    cleaned = re.sub(INVALID_TITLE_REGEX, "_", raw)
    if len(cleaned) <= 31:
        return cleaned
    suffix = hash(raw) & 0xFFFF
    prefix = cleaned[: 31 - len(str(suffix)) - 1]
    return f"{prefix}_{suffix}"


def write_doc_sheet(
    workbook: Workbook,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
) -> None:
    sheet = workbook.create_sheet(title=title)
    bold = Font(bold=True)
    header_cells = [WriteOnlyCell(sheet, value=h) for h in headers]
    for cell in header_cells:
        cell.font = bold
    sheet.append(header_cells)
    for row in rows:
        sheet.append([serialize_cell_value(value) for value in row])


def write_readme_sheet(
    workbook: Workbook,
    *,
    host: str,
    schemas: tuple[str, ...],
    table_rows: list[list[Any]],
) -> None:
    exported_at = datetime.now(UTC).isoformat()
    rows = [
        ["export_timestamp_utc", exported_at],
        ["connection_host", host],
        ["schemas", ", ".join(schemas)],
        ["table_count", len(table_rows)],
        ["notes", "Generated by neon-dump-2-xls"],
    ]
    write_doc_sheet(workbook, "_README", ["key", "value"], rows)


def stream_table_rows(
    conn: PgConnection,
    table: TableRef,
) -> tuple[list[str], Iterator[list[Any]]]:
    query = sql.SQL("SELECT * FROM {}.{}").format(
        sql.Identifier(table.schema_name),
        sql.Identifier(table.table_name),
    )
    cursor_name = f"export_{table.schema_name}_{table.table_name}"[:60]
    named_cur = conn.cursor(name=cursor_name)
    named_cur.itersize = BATCH_SIZE
    named_cur.execute(query)
    columns = [desc[0] for desc in named_cur.description or []]

    def row_iter() -> Iterator[list[Any]]:
        while True:
            batch = named_cur.fetchmany(BATCH_SIZE)
            if not batch:
                break
            for row in batch:
                yield [serialize_cell_value(value) for value in row]
        named_cur.close()

    return columns, row_iter()


def write_data_sheet(
    workbook: Workbook,
    conn: PgConnection,
    table: TableRef,
    used_names: set[str],
) -> str:
    base_name = sanitize_sheet_name(table.schema_name, table.table_name)
    sheet_name = base_name
    counter = 1
    while sheet_name in used_names:
        suffix = f"_{counter}"
        sheet_name = f"{base_name[: 31 - len(suffix)]}{suffix}"
        counter += 1
    used_names.add(sheet_name)

    sheet: Worksheet = workbook.create_sheet(title=sheet_name)
    columns, rows = stream_table_rows(conn, table)
    bold = Font(bold=True)
    header_cells = [WriteOnlyCell(sheet, value=col) for col in columns]
    for cell in header_cells:
        cell.font = bold
    sheet.append(header_cells)
    for row in rows:
        sheet.append([serialize_cell_value(value) for value in row])
    return sheet_name


def export_database_to_excel(
    database_url: str,
    output_path: str,
    *,
    schemas: tuple[str, ...] = DEFAULT_SCHEMAS,
    progress_callback: ProgressCallback | None = None,
) -> ExportSummary:
    exported_at = datetime.now(UTC)
    host = connection_host(database_url)

    with psycopg2.connect(database_url) as conn:
        conn.set_session(readonly=True, autocommit=False)
        tables = discover_tables(conn, schemas)
        inventory = fetch_table_inventory(conn, schemas)
        columns = fetch_columns(conn, schemas)
        foreign_keys = fetch_foreign_keys(conn, schemas)
        logical_config = load_logical_keys_config()
        logical_edges = explicit_logical_edges(logical_config)
        logical_rows = logical_key_rows(logical_config)
        mermaid = build_mermaid_er(foreign_keys, logical_edges)
        mermaid_sections = build_mermaid_er_sections(foreign_keys, logical_edges)

        workbook = Workbook(write_only=True)
        write_readme_sheet(
            workbook,
            host=host if isinstance(host, str) else "unknown-host",
            schemas=schemas,
            table_rows=inventory,
        )
        write_doc_sheet(
            workbook,
            "_Tables",
            [
                "schema",
                "table",
                "type",
                "approx_row_count",
                "column_count",
            ],
            inventory,
        )
        write_doc_sheet(
            workbook,
            "_Columns",
            [
                "schema",
                "table",
                "column",
                "type",
                "nullable",
                "default",
            ],
            columns,
        )
        write_doc_sheet(
            workbook,
            "_ForeignKeys",
            [
                "schema",
                "table",
                "column",
                "foreign_schema",
                "foreign_table",
                "foreign_column",
                "constraint_name",
            ],
            foreign_keys,
        )
        write_doc_sheet(
            workbook,
            "_LogicalKeys",
            [
                "source_schema",
                "source_table",
                "source_column",
                "target_schema",
                "target_table",
                "target_column",
                "link_type",
                "note",
            ],
            logical_rows,
        )
        write_doc_sheet(
            workbook,
            "_ER_Diagram",
            ["mermaid_source"],
            [[mermaid]],
        )

        used_names = set(DOC_SHEETS)
        total = len(tables)
        for index, table in enumerate(tables, start=1):
            if progress_callback:
                progress_callback(table.table_name, index, total)
            write_data_sheet(workbook, conn, table, used_names)

        workbook.save(output_path)

    diagram_path = Path(output_path).with_name(
        Path(output_path).stem + "_schema_diagram.html"
    )
    write_schema_diagram_html(
        diagram_path,
        mermaid_source=mermaid,
        mermaid_sections=mermaid_sections,
        host=host if isinstance(host, str) else "unknown-host",
        exported_at=exported_at,
        fk_count=len(foreign_keys),
        logical_count=len(logical_edges),
    )

    return ExportSummary(
        output_path=output_path,
        diagram_path=str(diagram_path),
        table_count=len(tables),
        sheet_count=len(used_names),
        exported_at=exported_at,
    )
